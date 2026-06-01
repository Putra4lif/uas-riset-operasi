"""Aplikasi web Metode Transportasi (Streamlit) — tampilan bergaya shadcn.

Menjalankan:
    pip install -r requirements.txt
    streamlit run app.py

Alur: input tabel interaktif (sumber, tujuan, biaya) -> solusi awal
(NWC / Least Cost / VAM) -> optimasi MODI/Stepping Stone -> langkah & solusi
optimal, diverifikasi otomatis dengan scipy. Seluruh ikon memakai Lucide (shadcn).
"""
from __future__ import annotations

import io
import math
import os

import pandas as pd
import streamlit as st

import ui
from solver import METHODS, load_case, solve, verify_with_scipy
from solver.model import TransportationProblem

HERE = os.path.dirname(os.path.abspath(__file__))
CONTOH = os.path.join(HERE, "cases", "semen.json")

st.set_page_config(page_title="Metode Transportasi", layout="wide")
ui.inject_css()


# --------------------------------------------------------------------------- #
# Util
# --------------------------------------------------------------------------- #
def fmt(x: float) -> str:
    """Format angka gaya Indonesia (titik ribuan). Kembalikan '—' bila kosong/NaN."""
    if not isinstance(x, (int, float)) or not math.isfinite(x):
        return "—"
    if abs(x - round(x)) < 1e-9:
        return f"{int(round(x)):,}".replace(",", ".")
    return f"{x:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


def to_num(x) -> float:
    """Konversi nilai sel tabel ke float. Sel kosong/teks -> NaN, ditolak oleh
    TransportationProblem.validate() dengan pesan ramah (bukan traceback)."""
    try:
        return float(x)
    except (TypeError, ValueError):
        return float("nan")


def muat_contoh() -> None:
    p = load_case(CONTOH)
    st.session_state.src_names = list(p.sources)
    st.session_state.src_supply = list(p.supply)
    st.session_state.dst_names = list(p.destinations)
    st.session_state.dst_demand = list(p.demand)
    st.session_state.cost = [list(r) for r in p.cost]


def init_state() -> None:
    if "src_names" not in st.session_state:
        muat_contoh()
        st.session_state.metode = "VAM"
        st.session_state.verifikasi = True


def unik(labels):
    seen, out = {}, []
    for x in labels:
        x = str(x) if x not in (None, "") else "?"
        if x in seen:
            seen[x] += 1
            out.append(f"{x} ({seen[x]})")
        else:
            seen[x] = 0
            out.append(x)
    return out


def _label_unik(base: str, dipakai) -> str:
    """Pilih label margin yang dijamin tak bentrok dengan nama sumber/tujuan pengguna.

    Mencegah bug: jika pengguna menamai tujuan "Kapasitas" atau sumber "Permintaan",
    label margin tidak boleh menimpa kolom/baris data tersebut.
    """
    label = base
    while label in dipakai:
        label += " "  # tambah spasi hingga unik
    return label


def resize_cost(m: int, n: int) -> None:
    cost = st.session_state.cost
    new = [[0.0] * n for _ in range(m)]
    for i in range(min(m, len(cost))):
        for j in range(min(n, len(cost[i]))):
            new[i][j] = cost[i][j]
    st.session_state.cost = new


# --------------------------------------------------------------------------- #
# Pembangun DataFrame (dipakai bersama oleh UI & ekspor)
# --------------------------------------------------------------------------- #
def alloc_matrix_df(problem: TransportationProblem, alloc) -> pd.DataFrame:
    """Matriks alokasi m x n + marjin Kapasitas (kolom) & Permintaan (baris).

    Label marjin dibuat unik agar tak menimpa data bila pengguna menamai
    tujuan "Kapasitas" atau sumber "Permintaan".
    """
    src = unik(problem.sources)
    dst = unik(problem.destinations)
    cap = _label_unik("Kapasitas", set(dst))
    dem = _label_unik("Permintaan", set(src))
    data = pd.DataFrame(
        [[alloc.alloc[i][j] for j in range(problem.n)] for i in range(problem.m)],
        index=src, columns=dst,
    )
    data[cap] = list(problem.supply)
    margin = pd.DataFrame(
        [list(problem.demand) + [problem.total_demand()]],
        index=[dem], columns=list(dst) + [cap],
    )
    return pd.concat([data, margin])


def route_df(problem: TransportationProblem, alloc,
             satuan_q: str = "", satuan_c: str = "") -> pd.DataFrame:
    """Rincian rute terpakai: Rute | Jumlah | Biaya/unit | Subtotal (memuat satuan)."""
    qh = f"Jumlah ({satuan_q})" if satuan_q else "Jumlah"
    ch = f"Biaya/unit ({satuan_c})" if satuan_c else "Biaya/unit"
    sh = f"Subtotal ({satuan_c})" if satuan_c else "Subtotal"
    baris = []
    for i in range(problem.m):
        for j in range(problem.n):
            q = alloc.alloc[i][j]
            if q > 1e-9:
                baris.append({
                    "Rute": f"{problem.sources[i]} → {problem.destinations[j]}",
                    qh: q,
                    ch: problem.cost[i][j],
                    sh: q * problem.cost[i][j],
                })
    return pd.DataFrame(baris)


def steps_df(problem: TransportationProblem, sol) -> pd.DataFrame:
    """Ringkasan tiap iterasi optimasi (untuk ditampilkan / diekspor)."""
    rows = []
    for it in sol.iterations:
        if it.optimal or it.entering is None:
            continue
        ei, ej = it.entering
        li, lj = it.leaving
        rows.append({
            "Iterasi": it.index,
            "Biaya saat ini": it.total_cost,
            "Sel masuk": f"{problem.sources[ei]} → {problem.destinations[ej]}",
            "Opportunity cost": it.reduced[it.entering],
            "Theta (digeser)": it.theta,
            "Sel keluar": f"{problem.sources[li]} → {problem.destinations[lj]}",
        })
    if not rows:
        rows = [{
            "Iterasi": "-", "Biaya saat ini": sol.optimal_cost,
            "Sel masuk": "(solusi awal sudah optimal)", "Opportunity cost": "-",
            "Theta (digeser)": "-", "Sel keluar": "-",
        }]
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------- #
# Ekspor, perbandingan metode, & diagram aliran
# --------------------------------------------------------------------------- #
def _style_ws(ws, title: str, tab_color: str, highlight_inner=None) -> None:
    """Format satu worksheet: judul, header berwarna, border, format angka, lebar kolom.

    Tabel ditulis mulai baris 2 (header) sehingga baris 1 dipakai untuk judul.
    highlight_inner=(m, n): sorot sel alokasi terpakai (>0) di blok m x n bagian dalam.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    n_col, n_row, header_row = ws.max_column, ws.max_row, 2
    thin = Side(style="thin", color="E4E4E7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Judul (baris 1), digabung selebar tabel.
    judul = ws.cell(row=1, column=1, value=title)
    judul.font = Font(size=13, bold=True, color="18181B")
    judul.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_col)
    ws.row_dimensions[1].height = 22

    # Baris header (latar gelap, teks putih).
    for c in range(1, n_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = PatternFill("solid", fgColor="18181B")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Sel data: border, format angka ribuan, rata kanan utk numerik.
    for r in range(header_row + 1, n_row + 1):
        for c in range(1, n_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.###"
                cell.alignment = Alignment(horizontal="right")

    # Sorot sel alokasi terpakai (>0) hijau.
    if highlight_inner:
        m, n = highlight_inner
        fill = PatternFill("solid", fgColor="DCFCE7")
        fnt = Font(color="166534", bold=True)
        for r in range(header_row + 1, header_row + 1 + m):   # baris sumber
            for c in range(2, 2 + n):                          # kolom tujuan
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, (int, float)) and cell.value > 1e-9:
                    cell.fill = fill
                    cell.font = fnt

    # Lebar kolom otomatis (abaikan baris judul agar kolom A tak melebar).
    for c in range(1, n_col + 1):
        longest = max(
            (len(str(ws.cell(row=r, column=c).value))
             for r in range(header_row, n_row + 1)
             if ws.cell(row=r, column=c).value is not None),
            default=10,
        )
        ws.column_dimensions[get_column_letter(c)].width = min(max(longest + 2, 11), 48)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color


def build_excel(sol, satuan_q: str = "", satuan_c: str = "") -> bytes:
    """Susun workbook .xlsx multi-sheet yang sudah diformat rapi (judul, header, dll.)."""
    p = sol.problem
    n_iter = len([it for it in sol.iterations if not it.optimal])
    ringkasan = pd.DataFrame({
        "Keterangan": [
            "Metode solusi awal", "Biaya solusi awal", "Biaya optimal",
            "Penghematan", "Jumlah iterasi", "Satuan jumlah", "Satuan biaya",
            "Penyeimbangan",
        ],
        "Nilai": [
            sol.initial_method, sol.initial_cost, sol.optimal_cost,
            sol.savings, n_iter, satuan_q or "-", satuan_c or "-",
            sol.balance_note,
        ],
    })
    judul_alloc = (f"Alokasi Optimal (dalam {satuan_q})" if satuan_q
                   else "Alokasi Optimal (unit dikirim)")
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        ringkasan.to_excel(writer, sheet_name="Ringkasan", index=False, startrow=1)
        alloc_matrix_df(p, sol.optimal).to_excel(writer, sheet_name="Alokasi Optimal", startrow=1)
        route_df(p, sol.optimal, satuan_q, satuan_c).to_excel(
            writer, sheet_name="Rincian Rute", index=False, startrow=1)
        steps_df(p, sol).to_excel(writer, sheet_name="Langkah", index=False, startrow=1)
        s = writer.sheets
        _style_ws(s["Ringkasan"], "Ringkasan Hasil — Metode Transportasi", "18181B")
        _style_ws(s["Alokasi Optimal"], judul_alloc, "16A34A", highlight_inner=(p.m, p.n))
        _style_ws(s["Rincian Rute"], "Rincian Biaya per Rute", "18181B")
        _style_ws(s["Langkah"], "Langkah Optimasi (MODI)", "18181B")
    return buf.getvalue()


def comparison_df(problem: TransportationProblem) -> pd.DataFrame:
    """Bandingkan ketiga metode solusi awal pada masalah yang sama."""
    rows = []
    for metode in METHODS:
        s = solve(problem, method=metode)
        rows.append({
            "Metode": metode,
            "Biaya awal": s.initial_cost,
            "Biaya optimal": s.optimal_cost,
            "Iterasi": len([it for it in s.iterations if not it.optimal]),
        })
    return pd.DataFrame(rows)


def _esc(s) -> str:
    return str(s).replace('"', "'")


def flow_diagram_dot(problem: TransportationProblem, alloc, satuan_q: str = "") -> str:
    """Bangun kode DOT (Graphviz) diagram aliran: sumber (kiri) → tujuan (kanan)."""
    sat = f" {satuan_q}" if satuan_q else ""
    out = [
        "digraph {",
        'rankdir=LR; bgcolor="transparent"; nodesep=0.28; ranksep=1.3;',
        'node [shape=box style="rounded,filled" fontname="Geist, sans-serif" fontsize=11 margin=0.14];',
        'edge [fontname="Geist, sans-serif" fontsize=10 color="#a1a1aa" penwidth=1.2];',
    ]
    for i in range(problem.m):
        out.append(
            f'S{i} [label="{_esc(problem.sources[i])}\\n({problem.supply[i]:g}{sat})" '
            'fillcolor="#f4f4f5" color="#e4e4e7"];'
        )
    for j in range(problem.n):
        out.append(
            f'D{j} [label="{_esc(problem.destinations[j])}\\n({problem.demand[j]:g}{sat})" '
            'fillcolor="#18181b" fontcolor="#fafafa" color="#18181b"];'
        )
    for i in range(problem.m):
        for j in range(problem.n):
            q = alloc.alloc[i][j]
            if q > 1e-9:
                out.append(f'S{i} -> D{j} [label="{q:g}"];')
    out.append("}")
    return "\n".join(out)


# --------------------------------------------------------------------------- #
# Render tabel
# --------------------------------------------------------------------------- #
def tabel_alokasi(problem: TransportationProblem, alloc, judul: str):
    df = alloc_matrix_df(problem, alloc)

    def sorot(val):
        return (
            "background-color:#dcfce7;color:#166534;font-weight:600"
            if isinstance(val, (int, float)) and val > 1e-9 else ""
        )

    st.caption(judul)
    # subset sudah mengecualikan baris/kolom margin (terakhir), jadi cukup sorot semua.
    sty = df.style.apply(
        lambda col: [sorot(v) for v in col],
        axis=0, subset=(df.index[:-1], df.columns[:-1]),
    ).format("{:g}")
    st.dataframe(sty, width="stretch")


def tabel_biaya_rute(problem: TransportationProblem, alloc,
                     satuan_q: str = "", satuan_c: str = ""):
    df = route_df(problem, alloc, satuan_q, satuan_c)
    fmt_map = {c: "{:g}" for c in df.columns if c != "Rute"}
    st.dataframe(df.style.format(fmt_map), width="stretch", hide_index=True)


def tampilkan_iterasi(problem: TransportationProblem, it):
    src = unik(problem.sources)
    dst = unik(problem.destinations)
    grid = []
    for i in range(problem.m):
        row = []
        for j in range(problem.n):
            row.append(f"{it.reduced[(i, j)]:+g}" if (i, j) in it.reduced else "basis")
        grid.append(row)
    df = pd.DataFrame(grid, index=src, columns=dst)
    st.caption("Opportunity cost cᵢⱼ − (uᵢ + vⱼ) untuk sel kosong (negatif = bisa diperbaiki):")
    st.dataframe(df, width="stretch")

    u_str = ", ".join(f"u{i+1}={'-' if it.u[i] is None else f'{it.u[i]:g}'}" for i in range(problem.m))
    v_str = ", ".join(f"v{j+1}={'-' if it.v[j] is None else f'{it.v[j]:g}'}" for j in range(problem.n))
    st.caption(f"Potensial: {u_str}  |  {v_str}")

    if it.entering is not None:
        ei, ej = it.entering
        li, lj = it.leaving
        lintasan = " → ".join(
            f"({problem.sources[a]}, {problem.destinations[b]})" for a, b in it.loop
        )
        st.markdown(
            f"- **Sel masuk:** ({problem.sources[ei]}, {problem.destinations[ej]}) "
            f"— opportunity cost {it.reduced[it.entering]:+g}\n"
            f"- **Lintasan tertutup (+, −, +, − …):** {lintasan}\n"
            f"- **θ (unit digeser):** {it.theta:g} — **sel keluar:** "
            f"({problem.sources[li]}, {problem.destinations[lj]})"
        )


# --------------------------------------------------------------------------- #
# Panel "Tentang" (edukatif)
# --------------------------------------------------------------------------- #
def render_about() -> None:
    """Penjelasan formal: definisi Metode Transportasi, tiap metode, & verifikasi scipy."""
    st.markdown(
        "**Metode Transportasi** merupakan salah satu kasus khusus *Linear Programming* yang "
        "bertujuan menentukan pola distribusi barang dari sejumlah **sumber** (misalnya pabrik "
        "atau gudang) ke sejumlah **tujuan** (misalnya kota atau wilayah pemasaran) sedemikian "
        "rupa sehingga **total biaya distribusi menjadi minimum**. Penyelesaian dilakukan dalam "
        "dua tahap, yaitu penentuan **solusi layak awal**, kemudian **pengujian dan perbaikan** "
        "hingga diperoleh solusi optimal."
    )
    st.markdown(
        "| Metode | Tahap | Prinsip kerja |\n"
        "|---|---|---|\n"
        "| **North-West Corner (NWC)** | Solusi awal | Pengalokasian dimulai dari sel pojok kiri-atas tanpa mempertimbangkan biaya. Paling sederhana, namun solusi awalnya umumnya belum efisien. |\n"
        "| **Least Cost** (Biaya Terkecil) | Solusi awal | Pengalokasian diprioritaskan pada sel dengan biaya satuan terkecil. Karena mempertimbangkan biaya, solusinya cenderung lebih baik daripada NWC. |\n"
        "| **VAM** (*Vogel's Approximation Method*) | Solusi awal | Menggunakan nilai *penalti*, yaitu selisih dua biaya terkecil pada setiap baris/kolom; alokasi dilakukan pada baris/kolom dengan penalti terbesar. Umumnya menghasilkan solusi yang paling mendekati optimal. |\n"
        "| **Stepping Stone** | Optimasi | Mengevaluasi setiap sel kosong melalui pembentukan lintasan tertutup untuk menghitung perubahan biaya, kemudian memindahkan alokasi. Bersifat intuitif dan visual. |\n"
        "| **MODI** (*Modified Distribution*) | Optimasi | Menghitung nilai potensial uᵢ dan vⱼ, kemudian menentukan *opportunity cost* cᵢⱼ − (uᵢ + vⱼ) secara matematis. Lebih efisien dibandingkan Stepping Stone. |"
    )
    st.markdown(
        "**Konsep pendukung.** *Keseimbangan:* total penawaran harus sama dengan total permintaan "
        "(Σ supply = Σ demand); apabila tidak seimbang, ditambahkan baris atau kolom *dummy* "
        "berbiaya nol. *Degenerasi:* solusi layak memerlukan tepat **m + n − 1** sel basis; "
        "apabila kurang, disisipkan sel beralokasi nol. *Optimalitas:* solusi dinyatakan optimal "
        "apabila seluruh *opportunity cost* sel kosong bernilai tak-negatif (≥ 0)."
    )
    ui.alert(
        "info", "badge-check", "Verifikasi dengan Solver Linear Programming (scipy)",
        "Masalah transportasi pada dasarnya merupakan bentuk khusus <b>Linear Programming (LP)</b>. "
        "Aplikasi ini menyelesaikan kembali persoalan yang sama menggunakan "
        "<code>scipy.optimize.linprog</code> sebagai pembanding independen. Apabila biaya optimal "
        "hasil perhitungan manual (MODI) identik dengan hasil solver LP tersebut, sistem "
        "menampilkan penanda <b>Terverifikasi</b> sebagai konfirmasi atas kebenaran algoritma."
    )


@st.dialog("Tentang Metode Transportasi", width="large")
def about_dialog() -> None:
    """Tampilkan penjelasan metode & verifikasi sebagai pop-up modal."""
    render_about()


# --------------------------------------------------------------------------- #
# Halaman
# --------------------------------------------------------------------------- #
init_state()

ui.hero(
    "truck",
    "Metode Transportasi",
    "Optimasi biaya distribusi: solusi awal NWC / Least Cost / VAM, dioptimalkan "
    "dengan MODI + Stepping Stone, lalu diverifikasi dengan solver LP scipy.",
    eyebrow="Riset Operasi",
)

with st.sidebar:
    ui.section("settings", "Pengaturan")
    st.session_state.metode = st.selectbox(
        "Metode solusi awal", ["NWC", "Least Cost", "VAM"],
        index=["NWC", "Least Cost", "VAM"].index(st.session_state.get("metode", "VAM")),
    )
    st.session_state.verifikasi = st.checkbox(
        "Verifikasi dengan scipy", value=st.session_state.get("verifikasi", True)
    )
    st.selectbox(
        "Satuan jumlah", ["ton", "kg", "kuintal", "unit", "liter", "karung", "dus", "m³"],
        index=0, key="satuan_jumlah", accept_new_options=True,
        help="Satuan untuk kapasitas, permintaan, & alokasi (boleh ketik sendiri).",
    )
    st.selectbox(
        "Satuan biaya", ["ribu Rp", "Rp", "juta Rp", "USD"],
        index=0, key="satuan_biaya", accept_new_options=True,
        help="Satuan biaya angkut per 1 satuan jumlah (boleh ketik sendiri).",
    )
    if st.button("Muat contoh PT Sentosa Beton", width="stretch"):
        muat_contoh()
        st.rerun()
    st.caption(
        "Ubah / tambah baris pada tabel Sumber & Tujuan, isi biaya angkut, "
        "lalu tekan Hitung Solusi Optimal."
    )
    open_about = st.button("Tentang metode & Solver LP", width="stretch", key="btn_about")

# Pop-up (modal) penjelasan metode & verifikasi scipy, dipicu dari tombol sidebar.
if open_about:
    about_dialog()

satuan_q = st.session_state.get("satuan_jumlah", "ton")
satuan_c = st.session_state.get("satuan_biaya", "ribu Rp")
sq_fmt = satuan_q.replace("%", "%%")  # escape agar aman dipakai di printf format kolom
sc_fmt = satuan_c.replace("%", "%%")

ui.section("package", "Sumber (asal) & kapasitas",
           "Langkah 1 — daftar pabrik/gudang & kapasitasnya")
src_df = st.data_editor(
    pd.DataFrame({"Sumber": st.session_state.src_names, "Kapasitas": st.session_state.src_supply}),
    num_rows="dynamic", width="stretch", key="ed_src",
    column_config={
        "Kapasitas": st.column_config.NumberColumn(
            "Kapasitas", format=f"%g {sq_fmt}", min_value=0,
            help=f"Kapasitas pasokan (dalam {satuan_q}).",
        ),
    },
)
src_df = src_df.dropna(how="all")
src_names = [str(x) for x in src_df["Sumber"].fillna("").tolist()]
src_supply = [to_num(x) for x in src_df["Kapasitas"].tolist()]

ui.section("map-pin", "Tujuan & permintaan",
           "Langkah 2 — daftar kota/wilayah & permintaannya")
dst_df = st.data_editor(
    pd.DataFrame({"Tujuan": st.session_state.dst_names, "Permintaan": st.session_state.dst_demand}),
    num_rows="dynamic", width="stretch", key="ed_dst",
    column_config={
        "Permintaan": st.column_config.NumberColumn(
            "Permintaan", format=f"%g {sq_fmt}", min_value=0,
            help=f"Permintaan tujuan (dalam {satuan_q}).",
        ),
    },
)
dst_df = dst_df.dropna(how="all")
dst_names = [str(x) for x in dst_df["Tujuan"].fillna("").tolist()]
dst_demand = [to_num(x) for x in dst_df["Permintaan"].tolist()]

st.session_state.src_names, st.session_state.src_supply = src_names, src_supply
st.session_state.dst_names, st.session_state.dst_demand = dst_names, dst_demand
m, n = len(src_names), len(dst_names)
resize_cost(m, n)

ui.section("grid", "Biaya angkut per unit",
           f"Langkah 3 — biaya tiap sumber → tujuan ({satuan_c} per {satuan_q})")
cost_df = pd.DataFrame(st.session_state.cost, index=unik(src_names), columns=unik(dst_names))
cost_cfg = {
    c: st.column_config.NumberColumn(c, format=f"%g {sc_fmt}", min_value=0)
    for c in cost_df.columns
}
cost_edit = st.data_editor(cost_df, width="stretch", key="ed_cost", column_config=cost_cfg)
st.session_state.cost = [[to_num(c) for c in row] for row in cost_edit.values.tolist()]

total_s, total_d = sum(src_supply), sum(dst_demand)
ui.stat_cards([
    {"label": "Total kapasitas", "value": f"{fmt(total_s)} {satuan_q}", "icon": "layers", "sub": "Penawaran (supply)"},
    {"label": "Total permintaan", "value": f"{fmt(total_d)} {satuan_q}", "icon": "package", "sub": "Permintaan (demand)"},
])
if abs(total_s - total_d) > 1e-9:
    ui.alert("info", "info", "Tidak seimbang",
             "Program akan menambahkan baris/kolom dummy (biaya 0) secara otomatis.")

st.write("")
if st.button("Hitung Solusi Optimal", type="primary", width="stretch"):
    try:
        problem = TransportationProblem(
            sources=src_names, destinations=dst_names,
            supply=src_supply, demand=dst_demand,
            cost=[[float(c) for c in row] for row in st.session_state.cost],
        )
        problem.validate()
        st.session_state.sol = solve(problem, method=st.session_state.metode)
        st.session_state.sol_error = ""
    except Exception as e:  # noqa: BLE001
        st.session_state.sol = None
        st.session_state.sol_error = str(e)

# Hasil disimpan di session_state agar TETAP tampil saat tombol unduh memicu rerun.
if st.session_state.get("sol_error"):
    ui.alert("destructive", "triangle-alert", "Input belum valid", st.session_state.sol_error)
elif st.session_state.get("sol") is not None:
    sol = st.session_state.sol

    st.divider()
    ui.alert("info", "info", "Penyeimbangan", sol.balance_note)

    n_iter = len([it for it in sol.iterations if not it.optimal])
    ui.stat_cards([
        {"label": "Biaya optimal", "value": fmt(sol.optimal_cost), "icon": "money",
         "sub": f"Total minimum ({satuan_c})"},
        {"label": "Biaya solusi awal", "value": fmt(sol.initial_cost), "icon": "calculator",
         "sub": f"{sol.initial_method} ({satuan_c})"},
        {"label": "Penghematan", "value": fmt(sol.savings), "icon": "trending-down",
         "accent": "green" if sol.savings > 1e-9 else "", "sub": f"Awal → optimal ({satuan_c})"},
        {"label": "Jumlah iterasi", "value": str(n_iter), "icon": "repeat", "sub": "MODI"},
    ])

    if st.session_state.verifikasi:
        ref = verify_with_scipy(sol.problem)
        if not ref["available"]:
            ui.alert("warning", "triangle-alert", "Verifikasi dilewati", ref["message"])
        elif ref["success"] and abs(ref["cost"] - sol.optimal_cost) < 1e-4:
            ui.alert("success", "badge-check", "Terverifikasi",
                     f"Hasil cocok dengan solver LP scipy (biaya = {fmt(ref['cost'])} {satuan_c}).")
        else:
            ui.alert("destructive", "triangle-alert", "Tidak cocok dengan scipy",
                     f"scipy = {ref['cost']}. Periksa kembali implementasi.")

    st.markdown("**Unduh hasil:**")
    dl1, dl2, _sp = st.columns([1, 1, 2])
    try:
        dl1.download_button(
            "Excel (.xlsx)", build_excel(sol, satuan_q, satuan_c),
            file_name="hasil_transportasi.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
    except Exception as e:  # noqa: BLE001
        dl1.caption(f"Excel butuh paket openpyxl ({e}).")
    dl2.download_button(
        "CSV (alokasi)",
        # utf-8-sig (BOM) agar Excel membaca karakter non-ASCII dengan benar;
        # float_format="%g" agar bilangan bulat tampil "50" bukan "50.0".
        alloc_matrix_df(sol.problem, sol.optimal).to_csv(float_format="%g").encode("utf-8-sig"),
        file_name="alokasi_optimal.csv", mime="text/csv", width="stretch",
    )

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        f"Solusi Awal ({sol.initial_method})", "Langkah Optimasi (MODI)",
        "Solusi Optimal", "Perbandingan Metode", "Diagram Aliran",
    ])
    with tab1:
        tabel_alokasi(sol.problem, sol.initial, f"Alokasi awal — {sol.initial_method} (dalam {satuan_q})")
        ui.metric("calculator", f"Total biaya solusi awal ({satuan_c})", fmt(sol.initial_cost))
    with tab2:
        langkah = [it for it in sol.iterations if not it.optimal]
        if not langkah:
            ui.alert("success", "circle-check", "Solusi awal sudah optimal",
                     "Tidak diperlukan iterasi perbaikan.")
        for it in langkah:
            with st.expander(f"Iterasi {it.index} — biaya saat ini: {fmt(it.total_cost)}"):
                tampilkan_iterasi(sol.problem, it)
        if langkah and sol.optimal_reached:
            ui.alert("success", "circle-check", "Optimal tercapai",
                     "Semua opportunity cost ≥ 0.")
    with tab3:
        tabel_alokasi(sol.problem, sol.optimal, f"Alokasi optimal (dalam {satuan_q})")
        tabel_biaya_rute(sol.problem, sol.optimal, satuan_q, satuan_c)
        ui.metric("money", f"Total biaya minimum ({satuan_c})", fmt(sol.optimal_cost), accent="green")
    with tab4:
        st.caption("Perbandingan ketiga metode solusi awal pada masalah yang sama "
                   "(semuanya menuju biaya optimal yang identik).")
        cmp = comparison_df(sol.original)
        st.dataframe(
            cmp.style.format({"Biaya awal": "{:g}", "Biaya optimal": "{:g}"}),
            width="stretch", hide_index=True,
        )
        st.caption(f"Biaya solusi awal tiap metode dalam {satuan_c} (makin rendah makin dekat ke optimal):")
        st.bar_chart(cmp.set_index("Metode")[["Biaya awal"]])
    with tab5:
        st.caption(f"Aliran distribusi optimal: sumber (kiri) → tujuan (kanan); "
                   f"angka pada panah = jumlah ({satuan_q}) yang dikirim.")
        st.graphviz_chart(flow_diagram_dot(sol.problem, sol.optimal, satuan_q))
